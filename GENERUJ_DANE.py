import os
import openpyxl
from tkinter import Tk, filedialog

def find_last_value_in_column(sheet, column):
    for row in range(sheet.max_row, 0, -1):
        cell_value = sheet.cell(row=row, column=column).value
        if cell_value is not None:
            return cell_value
    return None

def process_excel_files_in_directory(directory):
    all_data = []

    # Pobierz 21-24 znaki z nazwy katalogu
    dir_name_short = os.path.basename(directory)[20:24]

    for filename in os.listdir(directory):
        file_path = os.path.join(directory, filename)

        if os.path.isdir(file_path):
            # Jeżeli to katalog, przetwarzaj rekurencyjnie
            process_excel_files_in_directory(file_path)
        elif filename.endswith(".xlsx"):
            # Jeżeli to plik .xlsx, przetwarzaj
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)

                sheet = wb.active

                cell_b8 = sheet['B8'].value
                cell_a6 = sheet['A6'].value
                cell_d7 = sheet['D7'].value
                last_value_g_column = find_last_value_in_column(sheet, column=7)

                # Zamień kropki na przecinki
                if last_value_g_column is not None:
                    last_value_g_column = str(last_value_g_column).replace('.', ',')

                # Sprawdź, czy wartość nie jest None, zanim dodasz ją do listy
                last_value_g_column_str = str(last_value_g_column) if last_value_g_column is not None else "None"

                all_data.append(f"{filename}\t{cell_b8}\t{cell_a6}\t{cell_d7}\t{last_value_g_column_str}")

            except Exception as e:
                print(f"Błąd ładowania pliku {filename}: {e}")
                continue  # Przejdź do następnego pliku w przypadku błędu

    if all_data:
        # Pobierz 3 ostatnie znaki z nazwy katalogu
        dir_name_last_3_chars = os.path.basename(directory)[-3:]

        save_path = os.path.join(directory, f"{dir_name_short}_{dir_name_last_3_chars}_DANE.txt")

        with open(save_path, 'w') as txt_file:
            for line in all_data:
                txt_file.write(line + '\n')

def process_main_directory(main_directory):
    for root, dirs, files in os.walk(main_directory):
        for dir_name in dirs:
            dir_path = os.path.join(root, dir_name)
            process_excel_files_in_directory(dir_path)

if __name__ == "__main__":
    # Użyj okna dialogowego tkinter do wyboru katalogu głównego
    root = Tk()
    root.withdraw()  # Ukryj główne okno

    chosen_main_directory = filedialog.askdirectory(title="Wybierz główny katalog")

    if chosen_main_directory:
        process_main_directory(chosen_main_directory)

    # Zamknij okno tkinter po zakończeniu pracy
    root.destroy()
