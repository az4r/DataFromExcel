import os
import openpyxl
from tkinter import Tk, filedialog

def find_last_value_in_column(sheet, column):
    for row in range(sheet.max_row, 0, -1):
        cell_value = sheet.cell(row=row, column=column).value
        if cell_value is not None:
            return cell_value
    return None

def process_excel_files(directory):
    all_data = []

    for filename in os.listdir(directory):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(directory, filename)

            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active

            cell_b8 = sheet['B8'].value
            cell_a6 = sheet['A6'].value
            cell_d7 = sheet['D7'].value
            last_value_g_column = find_last_value_in_column(sheet, column=7)

            # Sprawdź, czy wartość nie jest None, zanim dodasz ją do listy
            last_value_g_column_str = str(last_value_g_column) if last_value_g_column is not None else "None"

            all_data.append(f"{filename}\t{cell_b8}\t{cell_a6}\t{cell_d7}\t{last_value_g_column_str}")

    save_path = os.path.join(directory, "dane.txt")
    with open(save_path, 'w') as txt_file:
        for line in all_data:
            txt_file.write(line + '\n')

if __name__ == "__main__":
    # Użyj okna dialogowego tkinter do wyboru katalogu
    root = Tk()
    root.withdraw()  # Ukryj główne okno

    chosen_directory = filedialog.askdirectory(title="Wybierz katalog z plikami Excela")

    if chosen_directory:
        process_excel_files(chosen_directory)
