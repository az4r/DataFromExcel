import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from tkinter import Tk, filedialog

def ustaw_szerokosc_kolumny(ws):
    # Ustawienia szerokości kolumn
    szerokosci_kolumn = [30, 30, 30, 115, 15, 15]

    for numer_kolumny, szerokosc in enumerate(szerokosci_kolumn, 1):
        litera_kolumny = get_column_letter(numer_kolumny)
        ws.column_dimensions[litera_kolumny].width = szerokosc

def ustaw_format_liczbowy(ws, numer_kolumny, format_liczbowy):
    # Ustaw format liczbowy dla danej kolumny
    litera_kolumny = get_column_letter(numer_kolumny)
    for wiersz in ws.iter_rows(min_col=numer_kolumny, max_col=numer_kolumny):
        for komorka in wiersz:
            komorka.number_format = format_liczbowy

def czy_liczba(wartosc):
    try:
        float(wartosc.replace(',', '.'))  # Zamień przecinek na kropkę przed konwersją
        return True
    except ValueError:
        return False

def txt_do_xlsx(sciezka_pliku_txt, katalog_wyjsciowy):
    # Otwórz plik tekstowy
    with open(sciezka_pliku_txt, 'r') as plik_txt:
        linie = plik_txt.readlines()

    # Utwórz nowy arkusz Excela
    wb = Workbook()
    arkusz = wb.active

    # Wypełnij arkusz danymi z pliku tekstowego
    for wiersz, linia in enumerate(linie, start=1):
        kolumny = linia.strip().split('\t')  # Zakładam, że dane w pliku tekstowym są rozdzielone tabulacją
        for kolumna, wartosc in enumerate(kolumny, start=1):
            # Sprawdź, czy wartość jest liczbą
            if kolumna == 2 and czy_liczba(wartosc):
                wartosc = round(float(wartosc.replace(',', '.')), 2)  # Zamień przecinek na kropkę przed konwersją
            # Zapisz wartość jako liczbę w kolumnie E
            if kolumna == 5:
                arkusz.cell(row=wiersz, column=kolumna, value=float(wartosc.replace(',', '.')))
            else:
                arkusz.cell(row=wiersz, column=kolumna, value=wartosc)

    # Ustaw szerokość kolumn
    ustaw_szerokosc_kolumny(arkusz)

    # Ustaw format liczbowy dla kolumny E
    ustaw_format_liczbowy(arkusz, numer_kolumny=5, format_liczbowy='0.00')

    # Dodaj formuły
    arkusz.cell(row=1, column=8, value="=E15+E16")
    arkusz.cell(row=2, column=8, value="=E11+E7+E8+E12")
    arkusz.cell(row=3, column=8, value="=E3+E4")
    arkusz.cell(row=4, column=8, value="=E19+E20")

    # Zapisz plik Excela z tą samą nazwą co plik tekstowy, ale z rozszerzeniem .xlsx
    nazwa_pliku_xlsx = os.path.join(katalog_wyjsciowy, os.path.splitext(os.path.basename(sciezka_pliku_txt))[0] + '.xlsx')
    wb.save(nazwa_pliku_xlsx)
    print(f"Plik {sciezka_pliku_txt} został przekonwertowany na {nazwa_pliku_xlsx}.")

def konwertuj_pliki_w_katalogu(katalog_wejsciowy):
    # Sprawdź, czy katalog istnieje
    if not os.path.exists(katalog_wejsciowy):
        print(f"Podany katalog {katalog_wejsciowy} nie istnieje.")
        return

    # Uzyskaj listę plików .txt w katalogu
    pliki_txt = [plik for plik in os.listdir(katalog_wejsciowy) if plik.endswith('.txt')]

    # Dla każdego pliku .txt, przekonwertuj go na plik .xlsx
    for plik_txt in pliki_txt:
        sciezka_pliku_txt = os.path.join(katalog_wejsciowy, plik_txt)
        txt_do_xlsx(sciezka_pliku_txt, katalog_wejsciowy)

if __name__ == "__main__":
    # Otwórz okno dialogowe wyboru katalogu
    root = Tk()
    root.withdraw()  # Ukryj główne okno
    katalog_wejsciowy = filedialog.askdirectory(title="Wybierz katalog z plikami .txt")

    # Jeśli użytkownik wybrał katalog, przekonwertuj pliki .txt na pliki .xlsx
    if katalog_wejsciowy:
        konwertuj_pliki_w_katalogu(katalog_wejsciowy)
        print("Konwersja zakończona.")
    else:
        print("Nie wybrano katalogu.")
